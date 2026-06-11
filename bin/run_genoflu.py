#!/usr/bin/env python
"""
run_genoflu.py — run USDA-VS GenoFLU on an assembled influenza-A FASTA,
capturing EVERY option used and writing ISO-aware provenance.

GenoFLU classifies North-American H5N1 2.3.4.4b HPAI viruses by BLASTing each
of the eight gene segments against a curated reference set and matching the
per-segment lineage pattern to a genotype key. Surveillance calls must be
defensible, so this wrapper:
  - records every command-line option, the percent-identity threshold, the
    reference-DB location + version, and tool versions used,
  - copies GenoFLU's native genotype output to stable filenames so the backend
    finds them without a timestamp glob,
  - parses the result into a structured genoflu_result.json (genotype + the
    per-segment lineage / %identity / mismatch / depth lists),
  - never raises — a non-zero GenoFLU exit is recorded, not fatal.

ISO / quality standards referenced in the provenance (for traceability):
  ISO 15189:2022 (medical lab quality: traceability, validation, version
  control, reporting), ISO/IEC 17025 (testing-lab competence; surveillance /
  veterinary), WOAH Terrestrial Manual Ch. 3.3.4 (avian influenza — reference
  characterization standard), WHO/WOAH/FAO H5 clade nomenclature (2.3.4.4b).
  The reportable per-segment metrics (% identity, mismatches, average depth)
  and the >=threshold% match rule are recorded per call so a genotype assignment
  can be verified independently.

Run standalone:
  python run_genoflu.py --fasta genome.fasta --outdir DIR --name SAMPLE \
      [--pident 98.0] [--genoflu-db /path/to/dependencies]
"""

import argparse
import csv
import json
import os
import re
import shutil
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

ISO_REFERENCES = [
    {"standard": "ISO 15189:2022", "scope": "Medical laboratory quality & competence (traceability, validation, version control, reporting)"},
    {"standard": "ISO/IEC 17025", "scope": "Testing-laboratory competence (surveillance / veterinary diagnostics)"},
    {"standard": "WOAH Terrestrial Manual 3.3.4", "scope": "Avian influenza — reference standard for detection & characterization"},
    {"standard": "WHO/WOAH/FAO H5 nomenclature", "scope": "Goose/Guangdong H5 clade naming (2.3.4.4b)"},
]

# The 8 influenza-A gene segments, in conventional order, for tidy reporting.
SEGMENT_ORDER = ["PB2", "PB1", "PA", "HA", "NP", "NA", "MP", "NS"]


def _now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def log(msg: str) -> None:
    print(msg, flush=True)


def _tool_version(cmd: List[str]) -> Optional[str]:
    try:
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        out = (proc.stdout or "").strip() or (proc.stderr or "").strip()
        return out.splitlines()[0].strip() if out else None
    except (FileNotFoundError, subprocess.SubprocessError, OSError):
        return None


def _find_genoflu() -> Optional[str]:
    """Locate the GenoFLU entry point on PATH (bioconda installs `genoflu.py`)."""
    for cand in ("genoflu.py", "genoflu"):
        which = shutil.which(cand)
        if which:
            return which
    return None


def _resolve_db(genoflu_exe: Optional[str], genoflu_db: Optional[str]) -> Dict[str, Any]:
    """Resolve the reference DB: explicit override, else the package's bundled
    `dependencies/` dir (relative to the installed genoflu.py). Returns
    {fastas, key, source, valid}."""
    info: Dict[str, Any] = {"fastas": None, "key": None, "source": None, "valid": False}

    def _populate(dep: Path, source: str) -> bool:
        fastas = dep / "fastas"
        key = dep / "genotype_key.xlsx"
        if fastas.is_dir() and key.is_file():
            info.update({"fastas": str(fastas), "key": str(key), "source": source, "valid": True})
            return True
        return False

    if genoflu_db:
        dep = Path(genoflu_db)
        # Accept either a `dependencies` dir or its parent.
        if _populate(dep, f"configured: {dep}"):
            return info
        if _populate(dep / "dependencies", f"configured: {dep}/dependencies"):
            return info
        log(f"WARNING: configured genoflu_db has no fastas/ + genotype_key.xlsx ({genoflu_db}); "
            "falling back to the conda package's bundled reference DB.")

    if genoflu_exe:
        # genoflu.py resolves dependencies at <script>/../dependencies
        script = Path(genoflu_exe).resolve()
        for dep in (script.parent.parent / "dependencies",
                    script.parent / "dependencies"):
            if _populate(dep, f"bundled: {dep}"):
                return info
    return info


def _db_version(db: Dict[str, Any]) -> Optional[str]:
    """Best-effort DB version: count of genotypes in the key + its mtime date.

    The genotype key has no formal version stamp, so we fingerprint it: the
    number of defined genotypes and the file's modification date uniquely
    identify the reference set used for a call."""
    key = db.get("key")
    if not key:
        return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(key, read_only=True)
        ws = wb.active
        rows = max(ws.max_row - 1, 0)  # minus header
        mtime = datetime.fromtimestamp(Path(key).stat().st_mtime).strftime("%Y-%m-%d")
        return f"{rows} genotypes (key dated {mtime})"
    except Exception:
        try:
            mtime = datetime.fromtimestamp(Path(key).stat().st_mtime).strftime("%Y-%m-%d")
            return f"key dated {mtime}"
        except OSError:
            return None


# ---------------------------------------------------------------------------
# Parse GenoFLU's native stats TSV into a structured result.
# Columns (GenoFLU >=1.05): sample, date, File Name, Metadata, Genotype,
#   "Genotype List Used, >={t}%", "Genotype Sample Title List",
#   "Genotype Percent Match List", "Genotype Mismatch List",
#   "Genotype Average Depth of Coverage List".
# ---------------------------------------------------------------------------
def _get(row: Dict[str, str], *needles: str) -> str:
    """Case-insensitive header lookup by substring (header wording shifts a bit
    across GenoFLU versions, e.g. the threshold embedded in a column name)."""
    for key, val in row.items():
        low = (key or "").lower()
        if all(n.lower() in low for n in needles):
            return (val or "").strip()
    return ""


def _split_list(s: str) -> List[str]:
    return [p.strip() for p in (s or "").split(",") if p.strip()]


def parse_genoflu_tsv(tsv_path: Path) -> Dict[str, Any]:
    """Parse the single-sample GenoFLU TSV into a structured dict."""
    result: Dict[str, Any] = {
        "genotype": None, "list_used": [], "segment_titles": [],
        "percent_match": [], "mismatch": [], "avg_depth": [],
        "segments": [], "segments_matched": 0, "complete": False, "raw": {},
    }
    if not tsv_path or not tsv_path.is_file():
        return result
    try:
        with tsv_path.open(newline="", encoding="utf-8", errors="replace") as fh:
            rows = list(csv.DictReader(fh, delimiter="\t"))
    except OSError:
        return result
    if not rows:
        return result
    row = rows[0]
    result["raw"] = {k: (v or "").strip() for k, v in row.items()}
    result["genotype"] = _get(row, "genotype") if "genotype" in {k.lower() for k in row} else None
    # `Genotype` is the bare column; prefer the exact one over substring hits.
    for k, v in row.items():
        if (k or "").strip().lower() == "genotype":
            result["genotype"] = (v or "").strip()
            break

    list_used = _split_list(_get(row, "list used"))
    titles = _split_list(_get(row, "sample title list"))
    pct = _split_list(_get(row, "percent match list"))
    mism = _split_list(_get(row, "mismatch list"))
    depth = _split_list(_get(row, "average depth"))
    result.update({
        "list_used": list_used, "segment_titles": titles,
        "percent_match": pct, "mismatch": mism, "avg_depth": depth,
    })

    # When run on a FASTA (no reads) GenoFLU emits a single non-per-segment
    # depth note like "Ran on FASTA - No Coverage Report"; only treat the depth
    # list as per-segment when it aligns 1:1 with the segments.
    depth_aligned = len(depth) == len(list_used)
    if depth and not depth_aligned:
        result["coverage_note"] = depth[0]

    # Build per-segment records by zipping the aligned lists. Each list_used
    # entry looks like "<gene>:<lineage>" (e.g. "HA:ea1").
    segs: List[Dict[str, Any]] = []
    for i, entry in enumerate(list_used):
        gene, _, lineage = entry.partition(":")
        segs.append({
            "segment": gene.strip() or entry,
            "lineage": lineage.strip() or None,
            "reference": titles[i] if i < len(titles) else None,
            "percent_identity": pct[i] if i < len(pct) else None,
            "mismatches": mism[i] if i < len(mism) else None,
            "avg_depth": depth[i] if depth_aligned else None,
        })

    def _seg_sort(s):
        seg = (s.get("segment") or "").upper()
        return SEGMENT_ORDER.index(seg) if seg in SEGMENT_ORDER else len(SEGMENT_ORDER)

    segs.sort(key=_seg_sort)
    result["segments"] = segs
    result["segments_matched"] = len(segs)
    gt = (result.get("genotype") or "")
    result["complete"] = bool(gt) and not gt.lower().startswith("not assigned")
    return result


# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------
def run(
    fasta: Path,
    outdir: Path,
    name: str,
    pident: float = 98.0,
    genoflu_db: Optional[str] = None,
    extra_provenance: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Run GenoFLU and write genoflu_genotype.{tsv,xlsx}, genoflu_result.json,
    and run_manifest.json. Returns the manifest dict."""
    outdir = Path(outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    genoflu_exe = _find_genoflu()
    db = _resolve_db(genoflu_exe, genoflu_db)

    if not genoflu_exe:
        log("ERROR: genoflu.py / genoflu not found on PATH.")
    else:
        log(f"GenoFLU: {genoflu_exe}")
    log(f"Reference DB: {db.get('source') or '(genoflu default — unresolved)'}")

    # GenoFLU opens the input FASTA by *basename relative to its CWD* (not the
    # path you pass), and writes its outputs to the CWD. So stage the input into
    # outdir and run GenoFLU there, referring to it by name. Flu genomes are
    # tiny, so a copy is cheap and keeps the analyzed input with its results.
    local_fasta = outdir / fasta.name
    try:
        if not local_fasta.exists() or fasta.resolve() != local_fasta.resolve():
            shutil.copyfile(fasta, local_fasta)
    except (OSError, RuntimeError) as exc:
        log(f"WARNING: could not stage input FASTA into the run dir: {exc}")
        local_fasta = fasta

    cmd: List[str] = []
    if genoflu_exe:
        cmd = [genoflu_exe, "-f", local_fasta.name, "-p", str(pident), "-n", name]
        if db.get("valid"):
            cmd += ["-i", db["fastas"], "-c", db["key"]]

    # Snapshot pre-existing *_stats.* so we can identify the new output.
    before = {p.name for p in outdir.glob("*_stats.*")}

    env = dict(os.environ)
    env.setdefault("TMPDIR", "/tmp")
    started = _now()
    rc = 0
    stderr_tail = ""
    if cmd:
        log(f"$ {' '.join(cmd)}")
        try:
            proc = subprocess.run(cmd, cwd=str(outdir), env=env, capture_output=True, text=True)
            rc = proc.returncode
            if proc.stdout:
                log(proc.stdout)
            if proc.stderr:
                print(proc.stderr, file=sys.stderr, flush=True)
                stderr_tail = "\n".join(proc.stderr.splitlines()[-20:])
        except (FileNotFoundError, OSError) as exc:
            rc = 127
            stderr_tail = f"GenoFLU failed to launch: {exc}"
            log(f"ERROR: {stderr_tail}")
    else:
        rc = 127
        stderr_tail = "genoflu executable not found"
    finished = _now()

    # Locate GenoFLU's native output (newest *_stats.tsv / *_stats.xlsx) and
    # copy to stable names the backend can find deterministically.
    native_tsv = _newest_new(outdir, "*_stats.tsv", before)
    native_xlsx = _newest_new(outdir, "*_stats.xlsx", before)
    stable_tsv = outdir / "genoflu_genotype.tsv"
    stable_xlsx = outdir / "genoflu_genotype.xlsx"
    if native_tsv:
        shutil.copyfile(native_tsv, stable_tsv)
    if native_xlsx:
        shutil.copyfile(native_xlsx, stable_xlsx)

    parsed = parse_genoflu_tsv(stable_tsv if stable_tsv.is_file() else native_tsv)
    (outdir / "genoflu_result.json").write_text(
        json.dumps(parsed, indent=2) + "\n", encoding="utf-8")
    if parsed.get("genotype"):
        log(f"GenoFLU genotype: {parsed['genotype']} "
            f"({parsed['segments_matched']} segments >= {pident}%)")

    manifest: Dict[str, Any] = {
        "tool": "GenoFLU",
        "sample": name,
        "input_fasta": str(fasta),
        "command": cmd,
        "started_at": started,
        "finished_at": finished,
        "return_code": rc,
        "options": {
            "pident_threshold": pident,
            "genoflu_db": db.get("source"),
            "fastas_dir": db.get("fastas"),
            "genotype_key": db.get("key"),
            "input_mode": "assembled FASTA (-f)",
        },
        "outputs": {
            "genotype_tsv": str(stable_tsv) if stable_tsv.is_file() else None,
            "genotype_xlsx": str(stable_xlsx) if stable_xlsx.is_file() else None,
            "result_json": str(outdir / "genoflu_result.json"),
        },
        "genotype": parsed.get("genotype"),
        "segments_matched": parsed.get("segments_matched"),
        "genotype_complete": parsed.get("complete"),
        "versions": {
            "genoflu": _tool_version([genoflu_exe, "-v"]) if genoflu_exe else None,
            "blastn": _tool_version(["blastn", "-version"]),
            "reference_db": _db_version(db),
        },
        "reportable_metrics_per_segment": [
            "percent identity (BLASTN)", "mismatches", "average depth of coverage",
        ],
        "thresholds_note": (
            f"A segment is counted as a match when its top BLASTN hit is "
            f">= {pident}% identity. A genotype is assigned only when all 8 "
            f"segments match the lineage pattern of a known genotype in the key; "
            f"otherwise the result is 'Not assigned' with the reason recorded."
        ),
        "iso_references": ISO_REFERENCES,
        "tmpdir": env.get("TMPDIR"),
        "stderr_tail": stderr_tail,
    }
    if extra_provenance:
        manifest.update(extra_provenance)
    (outdir / "run_manifest.json").write_text(
        json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    if rc != 0:
        log(f"WARNING: GenoFLU exited with code {rc}")
    return manifest


def _newest_new(outdir: Path, pattern: str, before: set) -> Optional[Path]:
    """Newest file matching `pattern` in outdir that wasn't there `before`.
    Falls back to the newest match if none are strictly new (e.g. a re-run)."""
    matches = sorted(outdir.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    new = [p for p in matches if p.name not in before]
    if new:
        return new[0]
    return matches[0] if matches else None


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Run GenoFLU with full provenance.")
    ap.add_argument("--fasta", type=Path, required=True)
    ap.add_argument("--outdir", type=Path, required=True)
    ap.add_argument("--name", required=True)
    ap.add_argument("--pident", type=float, default=98.0)
    ap.add_argument("--genoflu-db", default=None)
    args = ap.parse_args(argv)
    manifest = run(args.fasta, args.outdir, args.name,
                   pident=args.pident, genoflu_db=args.genoflu_db)
    return 0 if manifest.get("return_code", 1) == 0 else manifest.get("return_code", 1)


if __name__ == "__main__":
    sys.exit(main())
